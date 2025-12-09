# flake8: noqa: E501

import tkinter as tk
import re
import os
import logger
import threading
from tkinter import ttk, messagebox, font as tkFont
from tkcalendar import DateEntry
from db_manager import executionWithRs_query
from report.report_manager import ReportManager
from collections import defaultdict
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font
from LoadingPopup import LoadingPopupClass


logger.logger.info("[enquiryScreen] : Menu initiation")


def is_keyword_matched(keyword_words, desc_words, min_words_per_block=2, max_gap_per_block=3):
    found_positions = []
    keyword_words = keyword_words.lower().strip()
    desc_words = desc_words.lower().strip()

    # First strict full phrase check
    if keyword_words in desc_words:
        return True

    keyword_words = keyword_words.split()
    desc_words = desc_words.split()

    for kw in keyword_words:
        if kw in desc_words:
            pos = desc_words.index(kw)
            found_positions.append((kw, pos))

    if len(found_positions) == 0:
        return False

    # Sort by position
    found_positions.sort(key=lambda x: x[1])

    # Now check proximity block-wise
    blocks = []
    current_block = [found_positions[0]]

    for i in range(1, len(found_positions)):
        prev_word, prev_pos = found_positions[i-1]
        curr_word, curr_pos = found_positions[i]

        if curr_pos - prev_pos <= max_gap_per_block:
            current_block.append((curr_word, curr_pos))
        else:
            blocks.append(current_block)
            current_block = [(curr_word, curr_pos)]

    blocks.append(current_block)

    # Evaluate blocks
    for block in blocks:
        if len(block) >= min_words_per_block:
            return True  # Accept if any block has enough words close

    return False


class EnquiryScreen:

    def __init__(self, root, login_id):
        self.root = root
        # self.enquiry_window = tk.Toplevel(root)
        self.enquiry_window = tk.Toplevel(root)
        self.enquiry_window.title("Enquiry and Report Generation")

        # Register close cleanup
        self.enquiry_window.protocol("WM_DELETE_WINDOW", self.on_close)

        # === Create Canvas + Scrollbar
        self.main_canvas = tk.Canvas(self.enquiry_window,
                                     borderwidth=0, background="#f0f0f5")
        scrollbar = tk.Scrollbar(
            self.enquiry_window, orient="vertical", command=self.main_canvas.yview)
        self.scrollable_frame = tk.Frame(
            self.main_canvas, background="#f0f0f5")

        # === Configure scrollregion
        self.scrollable_frame.bind(
            "<Configure>",
            lambda e: self.main_canvas.configure(
                scrollregion=self.main_canvas.bbox("all")
            )
        )

        self.main_canvas.create_window(
            (0, 0), window=self.scrollable_frame, anchor="nw", tags="window")
        self.main_canvas.configure(yscrollcommand=scrollbar.set)

        def resize_scrollable_frame(event):
            self.main_canvas.itemconfig("window", width=event.width)

        self.main_canvas.bind("<Configure>", resize_scrollable_frame)

        # === Pack Canvas and Scrollbar
        self.main_canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # Optional: mouse wheel scrolling
        def _on_mousewheel(event):
            self.main_canvas.yview_scroll(
                int(-1 * (event.delta / 120)), "units")

        self.main_canvas.bind_all("<MouseWheel>", _on_mousewheel)

        screen_width = self.scrollable_frame.winfo_screenwidth()
        screen_height = self.scrollable_frame.winfo_screenheight()

        window_width = int(screen_width * 1)
        window_height = int(screen_height * 0.88)

        x_pos = int((screen_width - window_width) / 2)
        y_pos = int((screen_height - window_height) / 5)

        self.enquiry_window.geometry(
            f"{window_width}x{window_height}+{x_pos}+{y_pos}")
        self.enquiry_window.transient(self.root)  # Make it modal-like
        self.enquiry_window.grab_set()  # Prevent interaction with the main window
        self.enquiry_window.focus_force()  # Bring the focus to this window

        self.selection_states = []  # Stores ✓/✗ state per row
        self.next_summary_number = 1  # To avoid duplicated trn summ running number
        self.report_mgr = ReportManager()
        self.all_selected = True  # Default: tick all

        self.create_enquiry_header()
        self.create_enquiry_filters()
        self.create_layer2_results()
        self.create_action_buttons()
        self.create_layer3_summary()
        self.create_export_button()
        self.create_enquiry_footer()
        logger.logger.info("[enquiryScreen] : Screen establisted completely")

    def create_enquiry_header(self):
        logger.logger.info("[enquiryScreen] : Header - Deploying menu header section")

        header_frame = tk.Frame(self.scrollable_frame, bg="#4CAF50", height=60)
        header_frame.pack(fill=tk.X)

        title_label = tk.Label(header_frame, text="Enquiry and Report Generation",
                               font=("Helvetica", 18, "bold"), bg="#4CAF50", fg="white")
        title_label.pack(pady=10)
        breadcrumb_label = tk.Label(self.scrollable_frame, text="Home > Enquiry",
                                    font=("Helvetica", 12), bg="#f0f0f5", fg="#555")
        breadcrumb_label.pack(anchor="w", padx=20, pady=5)

    def create_enquiry_filters(self):
        logger.logger.info("[enquiryScreen] : Layer 1 - Deploying searching criteria section")

        self.customer_code_var = tk.StringVar()
        self.customer_name_var = tk.StringVar()
        self.customer_name_match_type = tk.StringVar(value="Equal")
        self.trx_desc_var = tk.StringVar()
        self.trx_desc_match_type = tk.StringVar(value="Equal")
        self.bank_name_var = tk.StringVar()
        self.trx_date_from_var = DateEntry(self.scrollable_frame)
        self.trx_date_to_var = DateEntry(self.scrollable_frame)
        # self.filter_entry_date_var = tk.StringVar(value="No")
        self.date_filter_type = tk.StringVar(value="Transaction Date")
        self.entry_date_from_var = DateEntry(self.scrollable_frame)
        self.entry_date_to_var = DateEntry(self.scrollable_frame)
        self.printed_status_var = tk.StringVar()
        self.agent_name_var = tk.StringVar()
        self.file_name_var = tk.StringVar()
        self.condition_var = tk.StringVar(value="All")

        filter_frame = tk.LabelFrame(
            self.scrollable_frame, text="Filters", bg="white", font=("Helvetica", 12))
        filter_frame_inner = tk.Frame(filter_frame, bg="white")
        filter_frame.pack(fill=tk.X, padx=20, pady=10)
        filter_frame_inner.pack(fill=tk.X)

        # Row 0 ────────────────────────────────────────────────────
        tk.Label(filter_frame_inner, text="Customer Code:", bg="white").grid(
            row=0, column=0, padx=10, pady=5, sticky='w')

        ttk.Entry(filter_frame_inner, textvariable=self.customer_code_var, width=23)\
            .grid(row=0, column=1, padx=10, pady=5, sticky='w')

        # self.customer_code_dropdown = ttk.Combobox(
        #     filter_frame_inner, textvariable=self.customer_code_var, state="readonly")
        # customer_code_data = executionWithRs_query(
        #     "SELECT DISTINCT VCH_CUST_CODE FROM TM_MST_CUSTOMER ORDER BY VCH_CUST_CODE")
        # customer_code_options = [
        #     "All"] + [row[0] for row in customer_code_data] if customer_code_data else ["All"]
        # self.customer_code_dropdown['values'] = customer_code_options
        # self.customer_code_var.set("All")
        # self.customer_code_dropdown.grid(
        #     row=0, column=1, padx=10, pady=5, sticky='e')

        tk.Label(filter_frame_inner, text="Customer Name:", bg="white").grid(
            row=0, column=2, padx=10, pady=5, sticky='w')
        name_frame = tk.Frame(filter_frame_inner, bg="white")
        name_frame.grid(row=0, column=3, padx=10, pady=5, sticky='e')
        ttk.Radiobutton(name_frame, text="Equal",
                        variable=self.customer_name_match_type, value="Equal").pack(side=tk.LEFT)
        ttk.Radiobutton(name_frame, text="Contain",
                        variable=self.customer_name_match_type, value="Contain").pack(side=tk.LEFT)
        ttk.Entry(name_frame, textvariable=self.customer_name_var,
                  width=20).pack(side=tk.LEFT, padx=5)

        tk.Label(filter_frame_inner, text="Transaction Desc:", bg="white").grid(
            row=0, column=4, padx=10, pady=5, sticky='w')
        desc_frame = tk.Frame(filter_frame_inner, bg="white")
        desc_frame.grid(row=0, column=5, padx=10, pady=5, sticky='e')
        ttk.Radiobutton(desc_frame, text="Equal",
                        variable=self.trx_desc_match_type, value="Equal").pack(side=tk.LEFT)
        ttk.Radiobutton(desc_frame, text="Contain",
                        variable=self.trx_desc_match_type, value="Contain").pack(side=tk.LEFT)
        ttk.Entry(desc_frame, textvariable=self.trx_desc_var,
                  width=20).pack(side=tk.LEFT, padx=5)

        tk.Label(filter_frame_inner, text="Bank Name:", bg="white").grid(
            row=0, column=6, padx=10, pady=5, sticky='w')
        self.bank_dropdown = ttk.Combobox(
            filter_frame_inner, textvariable=self.bank_name_var, state="readonly")
        bank_data = executionWithRs_query(
            "SELECT VCH_BANK_NAME FROM TM_MST_BANK ORDER BY VCH_BANK_NAME")
        bank_names = ["All"] + [row[0]
                                for row in bank_data] if bank_data else ["All"]
        self.bank_dropdown['values'] = bank_names
        self.bank_name_var.set("All")
        self.bank_dropdown.grid(row=0, column=7, padx=10, pady=5, sticky='e')

        # Row 1 ────────────────────────────────────────────────────
        tk.Label(filter_frame_inner, text="File Name:", bg="white").grid(row=1, column=0, padx=10, pady=5, sticky='w')
        self.file_dropdown = ttk.Combobox(filter_frame_inner, textvariable=self.file_name_var, state="readonly")
        file_data = executionWithRs_query(
            "SELECT DISTINCT VCH_FILE_NAME FROM TM_TRN_TRANSACTION ORDER BY VCH_FILE_NAME")
        file_names = ["All"] + [row[0]
                                for row in file_data] if file_data else ["All"]
        self.file_dropdown['values'] = file_names
        self.file_name_var.set("All")
        self.file_dropdown.grid(row=1, column=1, padx=10, pady=5, sticky='w')

        tk.Label(filter_frame_inner, text="Condition:", bg="white").grid(row=1, column=2, padx=10, pady=5, sticky='w')
        condition_dropdown = ttk.Combobox(
            filter_frame_inner,
            textvariable=self.condition_var,
            state="readonly",
            values=["All", "Blacklisted", "Blacklisted-PartialMatch",
                    "Suspicious", "Whitelist"]
        )
        condition_dropdown.grid(row=1, column=3, padx=10, pady=5, sticky='w')

        tk.Label(filter_frame_inner, text="Printed Status:", bg="white").grid(row=1, column=4, padx=10, pady=5, sticky='w')
        self.printed_status_var.set("All")
        printed_status_dropdown = ttk.Combobox(
            filter_frame_inner, textvariable=self.printed_status_var, state="readonly")
        printed_status_dropdown['values'] = ["All", "Yes", "No"]
        printed_status_dropdown.grid(row=1, column=5, padx=10, pady=5, sticky='w')

        tk.Label(filter_frame_inner, text="Agent Name:", bg="white").grid(row=1, column=6, padx=10, pady=5, sticky='w')
        self.agent_dropdown = ttk.Combobox(filter_frame_inner, textvariable=self.agent_name_var, state="readonly")
        agent_data = executionWithRs_query(
            "SELECT VCH_USER_NAME FROM TM_MST_USER ORDER BY VCH_USER_NAME")
        agent_names = ["All"] + [row[0]
                                 for row in agent_data] if agent_data else ["All"]
        self.agent_dropdown['values'] = agent_names
        self.agent_name_var.set("All")
        self.agent_dropdown.grid(row=1, column=7, padx=10, pady=5, sticky='w')

        # Row 2 ────────────────────────────────────────────────────
        tk.Label(filter_frame_inner, text="Filter by:", bg="white")\
            .grid(row=2, column=0, padx=10, pady=5, sticky="w")

        filter_type_frame = tk.Frame(filter_frame_inner, bg="white")
        filter_type_frame.grid(row=2, column=1, padx=10, pady=5, sticky="w")
        ttk.Radiobutton(filter_type_frame, text="Transaction Date", variable=self.date_filter_type,
                        value="Transaction Date", command=self.toggle_date_filter_type).pack(side=tk.LEFT)
        ttk.Radiobutton(filter_type_frame, text="Data Entry Date", variable=self.date_filter_type,
                        value="Data Entry Date", command=self.toggle_date_filter_type).pack(side=tk.LEFT)

        # -- Transaction Date --
        self.trx_date_from_label = tk.Label(filter_frame_inner, text="Transaction Date From:", bg="white")
        self.trx_date_from_label.grid(row=2, column=2, padx=10, pady=5, sticky='w')
        self.trx_date_from_var = DateEntry(filter_frame_inner)
        self.trx_date_from_var.grid(row=2, column=3, padx=10, pady=5, sticky='w')

        self.trx_date_to_label = tk.Label(filter_frame_inner, text="Transaction Date To:", bg="white")
        self.trx_date_to_label.grid(row=2, column=4, padx=10, pady=5, sticky='w')
        self.trx_date_to_var = DateEntry(filter_frame_inner)
        self.trx_date_to_var.grid(row=2, column=5, padx=10, pady=5, sticky='w')

        # -- Date Entry --
        self.entry_date_from_label = tk.Label(
            filter_frame_inner, text="Date Entry From:", bg="white", width=17)
        self.entry_date_from_label.grid(
            row=2, column=2, padx=10, pady=5, sticky="w")

        self.entry_date_from_var = DateEntry(filter_frame_inner)
        self.entry_date_from_var.grid(
            row=2, column=3, padx=10, pady=5, sticky="w")

        # -- Date Entry To --
        self.entry_date_to_label = tk.Label(
            filter_frame_inner, text="Date Entry To:", bg="white")
        self.entry_date_to_label.grid(
            row=2, column=4, padx=10, pady=5, sticky="w")

        self.entry_date_to_var = DateEntry(filter_frame_inner)
        self.entry_date_to_var.grid(
            row=2, column=5, padx=10, pady=5, sticky="w")

        # Row 3 ────────────────────────────────────────────────────
        # Button Frame
        btn_frame = tk.Frame(filter_frame_inner, bg="white")
        btn_frame.grid(row=3, column=0, columnspan=4, pady=10)
        ttk.Button(btn_frame, text="Search", command=self.search_enquiry).pack(
            side=tk.LEFT, padx=10)
        ttk.Button(btn_frame, text="Clear", command=self.clear_enquiry_filters).pack(
            side=tk.LEFT, padx=10)

        # make sure fields start hidden if default=No
        self.toggle_date_filter_type()

    def create_layer2_results(self):
        logger.logger.info("[enquiryScreen] : Layer 2 - Deploying search result section")

        # --- create a little header bar for results + legend
        results_header = tk.Frame(self.scrollable_frame, bg="#f0f0f5")
        results_header.pack(fill=tk.X, padx=20, pady=(10, 0))

        # Search Results label
        tk.Label(
            results_header,
            text="Search Results",
            bg="#f0f0f5",
            font=("Helvetica", 12, "bold")
        ).pack(side=tk.LEFT)

        # Legend container, right beside it
        legend_frame = tk.Frame(results_header, bg="#f0f0f5")
        legend_frame.pack(side=tk.LEFT, padx=(20, 0))

        # Red bullet
        tk.Label(
            legend_frame,
            text="● Blacklisted",
            font=("Helvetica", 9, "italic"),
            fg="red",
            bg="#f0f0f5"
        ).pack(side=tk.LEFT, padx=(0, 10))

        # Blue bullet
        tk.Label(
            legend_frame,
            text="● Blacklisted (Partial Match)",
            font=("Helvetica", 9, "italic"),
            fg="blue",
            bg="#f0f0f5"
        ).pack(side=tk.LEFT, padx=(0, 10))

        # Purple bullet
        tk.Label(
            legend_frame,
            text="● Suspicious",
            font=("Helvetica", 9, "italic"),
            fg="purple",
            bg="#f0f0f5"
        ).pack(side=tk.LEFT)

        # ✅ Fixed height frame to contain the table (approx. 10 rows)
        results_frame = tk.Frame(self.scrollable_frame)
        results_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=5)

        # ✅ Scrollbar
        scrollbar = ttk.Scrollbar(results_frame, orient=tk.VERTICAL)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # ✅ Treeview showing max 10 rows visually
        columns = (
            "✓/✗", "Customer Code", "Customer Name", "Transaction Description",
            "Target Audience", "Bank Name", "Credit Amount", "Debit Amount",
            "Transaction Date", "Data Entry Date", "Printed Status",
            "Agent Name", "File Name"
        )

        self.results_table = ttk.Treeview(
            results_frame,
            columns=columns,
            show="headings",
            height=10,
            yscrollcommand=scrollbar.set
        )

        style = ttk.Style()
        style.configure("Treeview", rowheight=20)

        self.results_table.tag_configure("evenrow", background="#f5f5f5")
        self.results_table.tag_configure("oddrow", background="#ffffff")
        self.results_table.tag_configure("redtext", background="#f5f5f5", foreground="red", font=("Helvetica", 10, "bold"))
        self.results_table.tag_configure("purpletext", background="#f5f5f5", foreground="purple")
        self.results_table.tag_configure("bluetext", background="#f5f5f5", foreground="blue")

        # ✅ Configure scrollbar command
        scrollbar.config(command=self.results_table.yview)

        for col in self.results_table["columns"]:
            self.results_table.heading(col, text=col)
            self.results_table.column(
                col, width=10 if col == "✓/✗" else 80, anchor="w")

        self.results_table.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        # Row Count Label
        self.result_count_label = tk.Label(
            self.scrollable_frame,
            text="",
            bg="#f0f0f5",
            font=("Helvetica", 10, "italic"),
            anchor="w"
        )
        self.result_count_label.pack(anchor="w", padx=20, pady=(0, 10))

        self.results_table.bind("<ButtonRelease-1>", self.toggle_selection)

        # Define column alignment one by one
        alignments = {
            "✓/✗": "center",
            "Customer Code": "center",
            "Customer Name": "w",
            "Transaction Description": "w",
            "Target Audience": "w",
            "Bank Name": "center",
            "Credit Amount": "center",
            "Debit Amount": "center",
            "Transaction Date": "center",
            "Data Entry Date": "center",
            "Printed Status": "center",
            "Agent Name": "w",
            "File Name": "w"
        }

        # Define column headings with sorting command
        for col in columns:
            anchor_style = alignments.get(col, "w")
            if col == "✓/✗":
                self.results_table.heading(
                    col, text=col, anchor="center",
                    command=self.toggle_all_selection)
            else:
                self.results_table.heading(
                    col,
                    text=col,
                    anchor="center",
                    command=lambda _col=col: self.sort_treeview_column(
                        self.results_table, _col, False)
                )
            self.results_table.column(
                col, anchor=anchor_style, width=120 if col != "✓/✗" else 60)

    def create_action_buttons(self):
        logger.logger.info("[enquiryScreen] : Layer 2.1 - Deploying search result buttons section")

        button_frame = tk.Frame(self.scrollable_frame, bg="#f0f0f5")
        button_frame.pack(pady=10)

        ttk.Button(button_frame, text="Add ➕", command=self.add_summary).pack(
            side=tk.LEFT, padx=10)
        ttk.Button(button_frame, text="Edit ✏️",
                   command=self.edit_summary).pack(side=tk.LEFT, padx=10)
        ttk.Button(button_frame, text="Delete ❌",
                   command=self.delete_summary).pack(side=tk.LEFT, padx=10)
        ttk.Button(button_frame, text="Reset 🔄",
                   command=self.reset_summary).pack(side=tk.LEFT, padx=10)

    def create_layer3_summary(self):
        logger.logger.info("[enquiryScreen] : Layer 2 - Deploying search result section")

        label = tk.Label(self.scrollable_frame, text="General Summary",
                         bg="#f0f0f5", font=("Helvetica", 12, "bold"))
        label.pack(anchor="w", padx=20)

        summary_frame = tk.Frame(self.scrollable_frame)
        # summary_frame.pack(fill=tk.X, padx=20, pady=5)
        summary_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=5)

        # ✅ Scrollbar
        scrollbar = ttk.Scrollbar(summary_frame, orient=tk.VERTICAL)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # self.summary_table = ttk.Treeview(summary_frame, columns=(
        #     "Summary Label", "Total Transactions", "Total Amount"), show="headings", height=5, yscrollcommand=scrollbar.set)
        self.summary_table = ttk.Treeview(summary_frame, columns=(
            "Summary Label", "Total Transactions", "Total Amount",
            "Total Customers", "Total Banks", "Transaction Date Range"
        ), show="headings", height=5, yscrollcommand=scrollbar.set)

        scrollbar.config(command=self.summary_table.yview)

        style = ttk.Style()
        style.configure("Treeview", rowheight=28)

        self.summary_table.tag_configure("evenrow", background="#f5f5f5")
        self.summary_table.tag_configure("oddrow", background="#ffffff")

        for col in self.summary_table["columns"]:
            self.summary_table.heading(col, text=col)
            self.summary_table.column(col, width=180, anchor="center")

        self.summary_table.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

    def create_export_button(self):
        logger.logger.info("[enquiryScreen] : Deploying the EXPORT (opt: Excel / Plaint Text) buttons")

        export_btn_frame = tk.Frame(self.scrollable_frame, bg="#f0f0f5")
        export_btn_frame.pack(fill=tk.X, padx=20, pady=(0, 10))

        ttk.Button(export_btn_frame, text="Export Plain Text", command=self.export_plain_text)\
            .pack(side=tk.RIGHT, padx=(5, 0))

        ttk.Button(export_btn_frame, text="Export Excel", command=self.export_excel)\
            .pack(side=tk.RIGHT, padx=(5, 0))

    def create_enquiry_footer(self):
        logger.logger.info("[enquiryScreen] : Footer - Deploying menu footer section")

        footer_frame = tk.Frame(self.scrollable_frame, bg="#4CAF50")
        footer_frame.pack(side=tk.BOTTOM, fill=tk.X, pady=10)

        label = tk.Label(footer_frame, text="Contact Helpdesk: Email: Euwin@example.com | Phone: +60 16-284 3121",
                         bg="#4CAF50", fg="white", font=("Helvetica", 10))
        label.pack()

    def toggle_selection(self, event):
        region = self.results_table.identify("region", event.x, event.y)
        if region == "cell":
            row_id = self.results_table.identify_row(event.y)
            col = self.results_table.identify_column(event.x)
            if col == "#1":  # only toggle if clicking the 'Select' column
                current = self.results_table.item(row_id)["values"]
                index = self.results_table.index(row_id)
                new_value = "✗" if current[0] == "✓" else "✓"
                self.selection_states[index] = (new_value == "✓")
                self.results_table.item(
                    row_id, values=[new_value] + current[1:])

    def search_enquiry(self):
        logger.logger.info("[enquiryScreen] : Executing the SEARCH operation, extract all transaction filter by parameter(s)")
        loading = LoadingPopupClass(
            self.enquiry_window, message="Searching... Please wait.")

        def threaded_search():
            try:
                self.perform_search_logic()
            finally:
                loading.close()

        threading.Thread(target=threaded_search).start()

    def perform_search_logic(self):
        # 1) clear existing
        self.results_table.delete(*self.results_table.get_children())
        self.selection_states.clear()

        # 2) collect filter values
        filters = {
            "customer_code":       self.customer_code_var.get(),
            "customer_name":       self.customer_name_var.get().strip(),
            "customer_name_match": self.customer_name_match_type.get(),
            "trx_desc":            self.trx_desc_var.get().strip(),
            "trx_desc_match":      self.trx_desc_match_type.get(),
            "bank_name":           self.bank_name_var.get(),
            "trx_date_from":       getattr(self.trx_date_from_var, "get_date", lambda: None)(),
            "trx_date_to":         getattr(self.trx_date_to_var,   "get_date", lambda: None)(),
            "printed_status":      self.printed_status_var.get(),
            "agent_name":          self.agent_name_var.get(),
            "file_name":           self.file_name_var.get(),
        }

        # Conditionally add entry date filters only if radio button is "Yes"
        if self.date_filter_type.get() == "Transaction Date":
            filters["trx_date_from"] = datetime.combine(
                self.trx_date_from_var.get_date(), datetime.min.time()) if self.trx_date_from_var.get() else None
            filters["trx_date_to"] = datetime.combine(
                self.trx_date_to_var.get_date(), datetime.min.time()) + timedelta(days=1) if self.trx_date_to_var.get() else None
            filters["entry_date_from"] = None
            filters["entry_date_to"] = None
        else:
            filters["entry_date_from"] = datetime.combine(
                self.entry_date_from_var.get_date(), datetime.min.time()) if self.entry_date_from_var.get() else None
            filters["entry_date_to"] = datetime.combine(
                self.entry_date_to_var.get_date(), datetime.min.time()) + timedelta(days=1) if self.entry_date_to_var.get() else None
            filters["trx_date_from"] = None
            filters["trx_date_to"] = None

        # 3) fetch real rows
        rows = self.report_mgr.fetch_transactions(filters)

        # 4) populate with check‐boxes and zebra tags
        if rows:
            for idx, row in enumerate(rows):
                tag = "evenrow" if idx % 2 == 0 else "oddrow"
                self.results_table.insert(
                    "", tk.END, values=["✓", *row], tags=(tag,))
                self.selection_states.append(True)

        # 5) Filter for blacklisted transaction
        self.filter_blacklisted()

        # 6) show the total number of rows 
        total_rows = len(self.results_table.get_children())
        self.result_count_label.config(text=f"Total Rows: {total_rows}")

    def clear_enquiry_filters(self):
        logger.logger.info("[enquiryScreen] : Executing the RESET operation, for searching criteria layer only")

        self.results_table.delete(*self.results_table.get_children())
        self.selection_states.clear()
        self.auto_fit_columns(self.results_table)

    def add_summary(self):
        logger.logger.info("[enquiryScreen] : Executing the ADD operation")

        # Collect only ticked rows (✓)
        selected_items = [
            item for item in self.results_table.get_children()
            if self.results_table.item(item, "values")[0] == "✓"
        ]

        if not selected_items:
            messagebox.showwarning("No Selection", "No records selected.")
            return

        # Prepare to collect data for summary
        self.selected_records = [self.results_table.item(
            item, "values") for item in selected_items]

        total_rows = len(self.selected_records)
        total_amount = 0
        customers = set()
        banks = set()
        dates = []

        for values in self.selected_records:
            customers.add(values[2])  # Customer Name
            banks.add(values[4])      # Bank Name
            dates.append(values[7])   # Transaction Date

            # Calculate total amount from credit/debit logic
            try:
                credit = float(values[5]) if values[5] else 0
                debit = float(values[6]) if values[6] else 0
                if credit != 0:
                    total_amount += credit
                else:
                    total_amount -= debit
            except:
                continue

        # To handle the action after the edit button from layer 3 - General Summary
        if not hasattr(self, 'selected_summaries'):
            self.selected_summaries = {}

        # Generate summary label
        if hasattr(self, 'editing_summary_label') and self.editing_summary_label:
            summary_label = self.editing_summary_label
            self.editing_summary_label = None
        else:
            summary_label = f"Transaction Summary #{self.next_summary_number}"
            self.next_summary_number += 1

        # Store selected records
        self.selected_summaries[summary_label] = self.selected_records.copy()

        # Format date range
        try:
            date_range = f"{min(dates)} until {max(dates)}"
        except:
            date_range = "-"

        index = len(self.summary_table.get_children())
        tag = "evenrow" if index % 2 == 0 else "oddrow"
        self.summary_table.insert(
            "", tk.END,
            values=(summary_label, total_rows, total_amount,
                    len(customers), len(banks), date_range),
            tags=(tag,)
        )

        self.results_table.delete(*self.results_table.get_children())
        self.selection_states.clear()
        self.result_count_label.config(text="Total Rows: 0")

    def edit_summary(self):
        logger.logger.info("[enquiryScreen] : Executing the EDIT operation")

        selected = self.summary_table.selection()
        if not selected:
            messagebox.showwarning(
                "No Selection", "Please select a summary row to edit.")
            return

        summary_label = self.summary_table.item(selected[0], "values")[0]
        self.editing_summary_label = summary_label  # store for reuse during update

        # ✅ Delete the summary row from table
        self.summary_table.delete(selected[0])

        # ✅ Restore the saved records to Layer 2
        records = self.selected_summaries.get(summary_label, [])
        if not records:
            messagebox.showwarning(
                "No Records", "No matching records found for this summary.")
            return

        self.results_table.delete(*self.results_table.get_children())
        self.selection_states.clear()

        for idx, record in enumerate(records):
            tag = "evenrow" if idx % 2 == 0 else "oddrow"
            self.results_table.insert("", tk.END, values=record, tags=(tag,))
            self.selection_states.append(True)  # mark all as selected again

        # ✅ Update Total Row count label
        total_rows = len(self.results_table.get_children())
        self.result_count_label.config(text=f"Total Rows: {total_rows}")

    def delete_summary(self):
        logger.logger.info("[enquiryScreen] : Executing the DELETE operation")

        selected = self.summary_table.selection()
        if not selected:
            messagebox.showwarning(
                "No Selection", "Please select a summary row to delete.")
            return
        # self.summary_table.delete(selected[0])
        summary_label = self.summary_table.item(selected[0], "values")[0]

        # ✅ Delete from Treeview
        self.summary_table.delete(selected[0])

        # ✅ Delete from stored summaries
        if hasattr(self, 'selected_summaries') and summary_label in self.selected_summaries:
            del self.selected_summaries[summary_label]

    def reset_summary(self):
        logger.logger.info("[enquiryScreen] : Executing the RESET operation, for entire enquiry module")

        confirm = messagebox.askyesno(
            "Reset", "Are you sure you want to clear all summary data?")
        if confirm:
            self.summary_table.delete(*self.summary_table.get_children())
            self.next_summary_number = 1  # reset the transaction summary number

    def on_close(self):
        self.main_canvas.unbind_all("<MouseWheel>")  # ✅ Unbind global scroll
        self.enquiry_window.destroy()                # ✅ Properly destroy window

    def auto_fit_columns(self, treeview):
        for col in treeview["columns"]:
            max_width = tkFont.Font().measure(col)  # Header width

            for row_id in treeview.get_children():
                cell_value = str(treeview.set(row_id, col))
                cell_width = tkFont.Font().measure(cell_value)
                if cell_width > max_width:
                    max_width = cell_width

            treeview.column(col, width=max_width + 20)  # +20 for padding

    def filter_blacklisted(self):
        # Fetch blacklist names
        bl_query = "SELECT VCH_BLACKLISTED_NAME FROM TM_MST_BLACKLISTED"
        bl_results = executionWithRs_query(bl_query)
        blacklist_names = [row[0].strip().lower()
                           for row in bl_results if row[0]] if bl_results else []

        # Fetch Suspicious names
        sp_query = "SELECT VCH_SUSPICIOUS_NAME FROM TM_MST_SUSPICIOUS"
        sp_results = executionWithRs_query(sp_query)
        suspicious_names = [row[0].strip().lower()
                            for row in sp_results if row[0]] if sp_results else []

        # Loop through Treeview rows
        for item_id in self.results_table.get_children():
            row_values = self.results_table.item(item_id)["values"]
            trx_desc = str(row_values[3]).strip().lower()

            # Replace '*' with space before full-match checks
            trx_desc_clean = trx_desc.replace("*", " ")
            tagged = False

            # --- Blacklist full match ---
            for bl_name in blacklist_names:
                if bl_name in trx_desc_clean:
                    self.results_table.item(item_id, tags=("redtext",))
                    tagged = True
                    break

            if not tagged:
                # --- Blacklist partial (word-based) match ---
                for bl_name in blacklist_names:
                    words = bl_name.split()
                    for word in words:
                        if len(word) > 1:
                            pattern = r'\b' + re.escape(word) + r'\b'
                            if re.search(pattern, trx_desc):
                                self.results_table.item(
                                    item_id, tags=("bluetext",))
                                tagged = True
                                break
                    if tagged:
                        break

            if not tagged:
                # --- Suspicious full match ---
                for sp_name in suspicious_names:
                    if sp_name in trx_desc_clean:
                        self.results_table.item(item_id, tags=("purpletext",))
                        break

    def toggle_all_selection(self):
        self.all_selected = not self.all_selected  # Flip state

        for idx, item_id in enumerate(self.results_table.get_children()):
            row_values = self.results_table.item(item_id)["values"]
            new_symbol = "✓" if self.all_selected else "✗"
            self.results_table.item(
                item_id, values=[new_symbol] + row_values[1:])
            self.selection_states[idx] = self.all_selected

    def sort_treeview_column(self, treeview, col, reverse):
        # Get data from Treeview
        data = [(treeview.set(k, col), k) for k in treeview.get_children("")]

        # Try numeric sort first, fall back to string sort
        try:
            data.sort(key=lambda t: float(
                str(t[0]).replace(",", "")), reverse=reverse)
        except ValueError:
            data.sort(key=lambda t: str(t[0]).lower(), reverse=reverse)

        # Reorder rows in Treeview
        for index, (val, k) in enumerate(data):
            treeview.move(k, "", index)

            # Zebra tagging
            tag = "evenrow" if index % 2 == 0 else "oddrow"
            treeview.item(k, tags=(tag,))

        # Toggle sort direction for next click
        treeview.heading(col, command=lambda: self.sort_treeview_column(
            treeview, col, not reverse))

    def export_plain_text(self):
        logger.logger.info("[enquiryScreen] : Executing the EXPORT operation into plain text")
        loading = LoadingPopupClass(
            self.enquiry_window, message="Exporting... Please wait.")

        def threaded_export():
            try:
                self.perform_plain_text_export()
            finally:
                loading.close()

        threading.Thread(target=threaded_export).start()

    def perform_plain_text_export(self):
        if not hasattr(self, 'selected_summaries') or not self.selected_summaries:
            messagebox.showinfo("No Data", "No summaries available to export.")
            return
        # Load blacklist and suspected keywords from DB
        bl_query = "SELECT VCH_BLACKLISTED_NAME FROM TM_MST_BLACKLISTED"
        bl_results = executionWithRs_query(bl_query)
        blacklist_names = [row[0].strip().lower()
                           for row in bl_results if row[0]] if bl_results else []

        sp_query = "SELECT VCH_SUSPICIOUS_NAME FROM TM_MST_SUSPICIOUS"
        sp_results = executionWithRs_query(sp_query)
        suspected_names = [row[0].strip().lower()
                           for row in sp_results if row[0]] if sp_results else []

        # Load bank display names
        bank_query = "SELECT VCH_BANK_NAME, VCH_BANK_DISPLAY_NM FROM TM_MST_BANK"
        bank_results = executionWithRs_query(bank_query)
        bank_display_map = {row[0].strip(): row[1].strip(
        ) for row in bank_results if row[0] and row[1]} if bank_results else {}

        grouped_output = {}
        final_agent_info = set()

        for records in self.selected_summaries.values():
            for row_values in records:
                values = row_values[1:]  # skip ✓/✗
                customer_code = values[0]
                customer_name = values[1]
                trx_desc = values[2]
                ner_name = values[3]
                bank = values[4]
                credit = values[5]
                debit = values[6]
                trx_date = values[7]
                agent_name = values[9]
                agent_id = values[10]

                final_agent_info.add((agent_name, agent_id))
                trx_desc_clean = trx_desc.lower().replace("*", " ")

                match_type = "Others"
                matched_key = "Others"

                for keyword in blacklist_names:
                    if is_keyword_matched(keyword, trx_desc_clean):
                        match_type = "Blacklisted"
                        matched_key = keyword.upper()
                        break
                else:
                    # --- Suspected full match ---
                    for keyword in suspected_names:
                        if keyword in trx_desc_clean:
                            match_type = "Suspected"
                            matched_key = keyword.upper()
                            break

                # Group by Customer
                cust_key = (customer_code, customer_name)
                if cust_key not in grouped_output:
                    grouped_output[cust_key] = {
                        "Blacklisted": defaultdict(list),
                        "Suspected": defaultdict(list),
                        "Others": []
                    }

                # Format transaction
                sign = "+"
                amount = credit
                if credit and float(credit.replace(",", "")) > 0:
                    sign = "+"
                    amount = credit
                elif debit and float(debit.replace(",", "")) > 0:
                    sign = "-"
                    amount = debit
                else:
                    amount = "0.00"

                try:
                    dt_obj = datetime.strptime(trx_date, "%Y-%m-%d %H:%M:%S")
                    trx_date_fmt = dt_obj.strftime(
                        "%d-%b-%y")  # e.g. 01-Dec-24
                except Exception:
                    trx_date_fmt = trx_date  # fallback to original if parsing fails

                # bank_display = bank_display_map.get(bank.strip(), bank)
                # formatted_line = f"{trx_date_fmt} | {bank_display} | RM {sign}{amount}"
                formatted_line = f"{trx_date_fmt} | RM {sign}{amount}"

                if match_type == "Others":
                    grouped_output[cust_key]["Others"].append(
                        (values[3], formatted_line))  # <-- add NER value here
                else:
                    grouped_output[cust_key][match_type][matched_key].append(
                        formatted_line)

        # Build output lines
        output_lines = []

        for (cust_code, cust_name), categories in grouped_output.items():
            output_lines.append(f"Customer ID: {cust_code}")
            output_lines.append(f"Customer: {cust_name}\n")

            # ✅ For Blacklisted
            if len(categories["Blacklisted"]) > 0:
                for category in ["Blacklisted"]:
                    output_lines.append("BLACKLISTED")
                    output_lines.append("===================================")
                    for keyword, lines in categories[category].items():
                        output_lines.append(f"{keyword}")
                        output_lines.append("- - - - - - - - - - - - - - - -")
                        output_lines.extend(lines)
                        output_lines.append("")

            # ✅ For Suspected
            if len(categories["Suspected"]) > 0:
                for category in ["Suspected"]:
                    output_lines.append("SUSPECTED")
                    output_lines.append("===================================")
                    for keyword, lines in categories[category].items():
                        output_lines.append(f"{keyword}")
                        output_lines.append("- - - - - - - - - - - - - - - -")
                        output_lines.extend(lines)
                        output_lines.append("")

            # Instead of Others, group by NER
            ner_group = defaultdict(list)
            for ner_value, trx_line in categories["Others"]:
                ner_group[ner_value].append(trx_line)

            # ✅ For Others
            if len(ner_group) > 0:
                output_lines.append("OTHERS")
                output_lines.append("===================================")
                for ner_name, lines in ner_group.items():
                    # Group by NER value (values[3])
                    output_lines.append(f"{ner_name}")
                    output_lines.append("- - - - - - - - - - - - - - - -")
                    output_lines.extend(lines)
                    output_lines.append("")

            output_lines.append("")  # spacing between customers

        for name, aid in final_agent_info:
            output_lines.append(f"Agent Name: {name}")

        # Create popup
        popup = tk.Toplevel(self.root)
        popup.title("Exported Plain Text")
        popup.geometry("1000x600")
        popup.transient(self.root)
        popup.grab_set()
        popup.focus_force()

        text_box = tk.Text(popup, wrap="word", font=("Courier", 10))
        text_box.insert("1.0", "\n".join(output_lines))
        text_box.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        def copy_to_clipboard():
            popup.clipboard_clear()
            popup.clipboard_append(text_box.get("1.0", tk.END))
            messagebox.showinfo("Copied", "Text copied to clipboard.")

        ttk.Button(popup, text="Copy", command=copy_to_clipboard).pack(pady=5)

        # ✅ Cleanup after export
        # self.summary_table.delete(*self.summary_table.get_children())
        # self.selected_summaries = {}
        # self.next_summary_number = 1
        # self.results_table.delete(*self.results_table.get_children())
        # self.selection_states.clear()

    def export_excel(self):
        logger.logger.info("[enquiryScreen] : Executing the EXPORT operation into Excel")
        loading = LoadingPopupClass(
            self.enquiry_window, message="Exporting... Please wait.")

        def threaded_export():
            try:
                self.perform_excel_export()
            finally:
                loading.close()

        threading.Thread(target=threaded_export).start()

    def perform_excel_export(self):
        if not hasattr(self, 'selected_summaries') or not self.selected_summaries:
            messagebox.showinfo("No Data", "No summaries available to export.")
            return

        # Save path
        filename = f"Transaction_Export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(os.path.expanduser("~"), "Downloads", filename)

        wb = Workbook()
        ws = wb.active
        ws.title = "Transactions"

        # Styles
        bold_font = Font(bold=True)

        # Bank display mapping
        bank_query = "SELECT VCH_BANK_NAME, VCH_BANK_DISPLAY_NM FROM TM_MST_BANK"
        bank_results = executionWithRs_query(bank_query)
        bank_display_map = {row[0].strip(): row[1].strip(
        ) for row in bank_results if row[0] and row[1]} if bank_results else {}

        # Keywords
        bl_query = "SELECT VCH_BLACKLISTED_NAME FROM TM_MST_BLACKLISTED"
        blacklist = [r[0].strip().lower()
                     for r in executionWithRs_query(bl_query) or [] if r[0]]

        sp_query = "SELECT VCH_SUSPICIOUS_NAME FROM TM_MST_SUSPICIOUS"
        suspected = [r[0].strip().lower()
                     for r in executionWithRs_query(sp_query) or [] if r[0]]

        # Grouped output: (Customer Code, Name) → list of rows
        grouped_output = defaultdict(list)

        for records in self.selected_summaries.values():
            for row in records:
                values = row[1:]  # skip ✓/✗
                cust_id = values[0]
                cust_name = values[1]
                trx_desc_raw = values[2]
                trx_desc = trx_desc_raw.lower().replace("*", " ")
                trx_ner = values[3]
                bank = values[4]
                credit = values[5]
                debit = values[6]
                trx_date = values[7]
                entry_date = values[8]
                printed = values[9]
                agent_name = values[10]
                file_name = values[11]

                # Match logic
                category = "Others"
                keyword = trx_ner
                # for kw in blacklist:
                #     if kw in trx_desc or any(re.search(rf'\b{re.escape(w)}\b', trx_desc) for w in kw.split()):
                #         category = "Blacklisted"
                #         keyword = kw.upper()
                #         break
                for kw in blacklist:
                    if is_keyword_matched(kw, trx_desc):
                        category = "Blacklisted"
                        keyword = kw.upper()
                        break
                else:
                    for kw in suspected:
                        if kw in trx_desc:
                            category = "Suspected"
                            keyword = kw.upper()
                            break

                # Format date
                try:
                    trx_dt = datetime.strptime(trx_date, "%Y-%m-%d %H:%M:%S")
                    trx_fmt = trx_dt.strftime("%d-%b-%y")
                except ValueError:
                    trx_fmt = trx_date

                # Bank display
                bank_display = bank_display_map.get(bank.strip(), bank)

                # Store in grouped structure
                cust_key = (cust_id, cust_name)
                grouped_output[cust_key].append([
                    keyword, category, trx_fmt, trx_desc_raw,
                    trx_ner, bank_display, credit, debit,
                    entry_date, printed, agent_name, file_name
                ])

        # Output to sheet
        for (cust_id, cust_name), rows in grouped_output.items():
            ws.append([f"Customer ID: {cust_id}"])
            ws.append([f"Customer Name: {cust_name}"])
            ws.append([])

            # Column headers
            headers = [
                "Category Keyword", "Category", "Transaction Date",
                "Transaction Description", "Target Audience",
                "Bank Display Name", "Credit Amount", "Debit Amount",
                "Data Entry Date", "Printed Status", "Agent Name", "File Name"
            ]
            ws.append(headers)
            for cell in ws[ws.max_row]:
                cell.font = bold_font

            for r in rows:
                ws.append(r)

            ws.append([])  # spacing

        # Save
        try:
            wb.save(filepath)
            messagebox.showinfo("Success", f"Excel exported to:\n{filepath}")
        except Exception as e:
            messagebox.showerror(
                "Export Failed", f"Could not export Excel:\n{e}")

        # ✅ Cleanup after export
        # self.summary_table.delete(*self.summary_table.get_children())
        # self.selected_summaries = {}
        # self.next_summary_number = 1
        # self.results_table.delete(*self.results_table.get_children())
        # self.selection_states.clear()

    def toggle_date_filter_type(self):
        if self.date_filter_type.get() == "Transaction Date":
            self.trx_date_from_label.grid()
            self.trx_date_from_var.grid()
            self.trx_date_to_label.grid()
            self.trx_date_to_var.grid()
            self.entry_date_from_label.grid_remove()
            self.entry_date_from_var.grid_remove()
            self.entry_date_to_label.grid_remove()
            self.entry_date_to_var.grid_remove()
        else:
            self.trx_date_from_label.grid_remove()
            self.trx_date_from_var.grid_remove()
            self.trx_date_to_label.grid_remove()
            self.trx_date_to_var.grid_remove()
            self.entry_date_from_label.grid()
            self.entry_date_from_var.grid()
            self.entry_date_to_label.grid()
            self.entry_date_to_var.grid()
